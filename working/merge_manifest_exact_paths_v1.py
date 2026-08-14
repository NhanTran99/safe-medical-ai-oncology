from pathlib import Path
from openpyxl import load_workbook
from shutil import copy2
from datetime import datetime

REPO = Path(r"D:/GIT hub/safe-medical-ai-oncology")
MANIFEST = REPO / "working" / "POPULATION_PACKAGE_INTEGRATION_MANIFEST_completed.xlsx"
MAPPING = REPO / "working" / "PP_EXACT_REPOSITORY_PATH_MAPPING.tsv"
OUTPUT = REPO / "working" / "POPULATION_PACKAGE_INTEGRATION_MANIFEST_exact_paths_v1.xlsx"
BACKUP = REPO / "working" / "POPULATION_PACKAGE_INTEGRATION_MANIFEST_completed.backup.xlsx"

if not MANIFEST.exists():
    raise FileNotFoundError(f"Manifest not found: {MANIFEST}")
if not MAPPING.exists():
    raise FileNotFoundError(f"Mapping not found: {MAPPING}")

# Load exact Git-derived mapping.
mapping = {}
with MAPPING.open("r", encoding="utf-8-sig", newline="") as f:
    for line_no, raw in enumerate(f, 1):
        line = raw.rstrip("\r\n")
        if not line:
            continue
        parts = line.split("\t", 1)
        if len(parts) != 2:
            # The user's current TSV display may visually collapse tabs;
            # accept one-or-more whitespace only if the path begins with /.
            parts = line.split(None, 1)
        if len(parts) != 2:
            raise ValueError(f"Invalid mapping line {line_no}: {raw!r}")
        pp_id, repo_path = parts[0].strip(), parts[1].strip()
        if pp_id in mapping:
            raise ValueError(f"Duplicate PP ID in mapping: {pp_id}")
        if not repo_path.startswith("/") or not repo_path.endswith("/"):
            raise ValueError(f"Non-canonical repository path for {pp_id}: {repo_path}")
        mapping[pp_id] = repo_path

if len(mapping) != 239:
    raise ValueError(f"Expected 239 mapping rows; found {len(mapping)}")

# Load workbook without changing unrelated cells/sheets.
wb = load_workbook(MANIFEST)
if "Sheet1" not in wb.sheetnames:
    raise ValueError(f"Expected Sheet1; found {wb.sheetnames}")
ws = wb["Sheet1"]

headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
required = ["PP ID", "Repository Path", "Status", "Notes"]
missing = [h for h in required if h not in headers]
if missing:
    raise ValueError(f"Missing required columns: {missing}")

pp_col = headers["PP ID"]
path_col = headers["Repository Path"]
status_col = headers["Status"]
notes_col = headers["Notes"]

rows_by_id = {}
for r in range(2, ws.max_row + 1):
    pp_id = ws.cell(r, pp_col).value
    if isinstance(pp_id, str) and pp_id.startswith("PP-"):
        if pp_id in rows_by_id:
            raise ValueError(f"Duplicate PP ID in Manifest: {pp_id}")
        rows_by_id[pp_id] = r

if len(rows_by_id) != 239:
    raise ValueError(f"Expected 239 PP rows in Manifest; found {len(rows_by_id)}")

missing_in_manifest = sorted(set(mapping) - set(rows_by_id))
missing_in_mapping = sorted(set(rows_by_id) - set(mapping))
if missing_in_manifest or missing_in_mapping:
    raise ValueError(
        f"PP ID mismatch. Missing in Manifest: {missing_in_manifest}; "
        f"missing in mapping: {missing_in_mapping}"
    )

# Backup original before writing output.
if not BACKUP.exists():
    copy2(MANIFEST, BACKUP)

# Apply ONLY the intended integration-state edits.
for pp_id, r in rows_by_id.items():
    ws.cell(r, path_col).value = mapping[pp_id]
    ws.cell(r, status_col).value = "PENDING"
    ws.cell(r, notes_col).value = "GOLD aggregate verification = PENDING"

# Re-open/save validation.
wb.save(OUTPUT)
check = load_workbook(OUTPUT, data_only=False)
ws2 = check["Sheet1"]
headers2 = {ws2.cell(1, c).value: c for c in range(1, ws2.max_column + 1)}

errors = []
for pp_id, r in rows_by_id.items():
    rr = r
    if ws2.cell(rr, headers2["Repository Path"]).value != mapping[pp_id]:
        errors.append(f"{pp_id}: Repository Path mismatch after save")
    if ws2.cell(rr, headers2["Status"]).value != "PENDING":
        errors.append(f"{pp_id}: Status != PENDING")
    if ws2.cell(rr, headers2["Notes"]).value != "GOLD aggregate verification = PENDING":
        errors.append(f"{pp_id}: Notes mismatch")

if errors:
    raise RuntimeError("POST-SAVE VALIDATION FAILED:\n" + "\n".join(errors[:20]))

print("MERGE PASS")
print(f"PP rows updated: {len(rows_by_id)}")
print("Repository Path: 239/239 exact Git-derived mappings")
print("Status: 239/239 PENDING")
print("Notes: 239/239 GOLD aggregate verification = PENDING")
print(f"Backup: {BACKUP}")
print(f"Output: {OUTPUT}")
