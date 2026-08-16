#!/usr/bin/env python3
"""Regenerate the runtime Evaluation Case Manifest projection — Track 2.

Reads the frozen, authoritative
`08_Development/implementation/data/PHASE6_STAGE2_EVALUATION_CASE_MANIFEST_v1.0_FROZEN.xlsx`
and deterministically (re)writes
`08_Development/implementation/data/evaluation_case_manifest_projection.json`
— a runtime-consumable projection containing only the fields the runtime
needs: `case_id`, `population_id`, `expected_primary_artifact_type`
(consumed by `EvaluationCaseResolver`), plus `pp_title` and
`controlled_question` (consumed only by the Chat UI's controlled
navigation catalog — see `api/main.py`'s `_load_navigation_catalog()` —
never used for case *resolution*, only for non-clinical display labels
and the pre-approved question-starter text). The projection also carries
the source manifest's identity, version, and SHA-256 hash for
traceability.

This script is a dev-time tool only. It is never imported by the FastAPI
runtime and requires `openpyxl` (a dev-only dependency — see
`pyproject.toml`'s `[project.optional-dependencies].dev` — not a runtime
dependency; production code never parses XLSX).

The generated projection is a derived, regenerable binding, never an
independent source of truth: any change to the authoritative frozen
manifest must flow through re-running this script, never through
hand-editing the projection JSON directly. Before writing, this script
validates the frozen source against the invariants the runtime relies on
(239 cases, unique/complete 1:1 EC<->PP mapping, a valid known artifact
type, no missing required fields) and fails loudly rather than emitting a
partial or inconsistent projection.

Usage:
    uv run --extra dev python 08_Development/implementation/scripts/generate_evaluation_case_manifest_projection.py
"""

from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path

import openpyxl

_HERE = Path(__file__).resolve().parent
_DATA_DIR = _HERE.parent / "data"
_MANIFEST_PATH = _DATA_DIR / "PHASE6_STAGE2_EVALUATION_CASE_MANIFEST_v1.0_FROZEN.xlsx"
_PROJECTION_PATH = _DATA_DIR / "evaluation_case_manifest_projection.json"

_SOURCE_MANIFEST_NAME = "PHASE6_STAGE2_EVALUATION_CASE_MANIFEST_v1.0_FROZEN.xlsx"
_SOURCE_MANIFEST_VERSION = "1.0-FROZEN"
_SHEET_NAME = "Evaluation Manifest"

_EXPECTED_CASE_COUNT = 239
_CASE_ID_PATTERN = re.compile(r"^EC-\d{4}$")
_POPULATION_ID_PATTERN = re.compile(r"^PP-\d{4}$")
_KNOWN_ARTIFACT_TYPES = {"CKO", "KNOWLEDGE_PASSPORT", "PRIMARY_EVIDENCE_PACKAGE", "QA_REPORT"}


def _validate(cases: list[dict[str, str]]) -> None:
    """Fail loudly if the frozen source violates a runtime invariant.

    Deliberately minimal: this checks only what the runtime actually
    relies on (uniqueness, 1:1 mapping, known artifact type, required
    fields present) -- it is not a general-purpose manifest QA engine.
    """
    errors: list[str] = []

    if len(cases) != _EXPECTED_CASE_COUNT:
        errors.append(f"expected {_EXPECTED_CASE_COUNT} cases, found {len(cases)}")

    case_ids = [c["case_id"] for c in cases]
    population_ids = [c["population_id"] for c in cases]

    if len(set(case_ids)) != len(case_ids):
        errors.append("duplicate Case ID values found")
    if len(set(population_ids)) != len(population_ids):
        errors.append("duplicate PP ID values found")

    for case in cases:
        case_id = case["case_id"]
        population_id = case["population_id"]
        artifact_type = case["expected_primary_artifact_type"]
        pp_title = case["pp_title"]
        controlled_question = case["controlled_question"]

        if not case_id or not _CASE_ID_PATTERN.match(str(case_id)):
            errors.append(f"malformed Case ID: {case_id!r}")
            continue
        if not population_id or not _POPULATION_ID_PATTERN.match(str(population_id)):
            errors.append(f"malformed PP ID: {population_id!r} (case {case_id})")
        # 1:1 mapping: EC-NNNN must map to PP-NNNN exactly.
        if population_id != f"PP-{case_id.split('-')[1]}":
            errors.append(f"EC/PP mismatch: {case_id} -> {population_id}")
        if artifact_type not in _KNOWN_ARTIFACT_TYPES:
            errors.append(f"unknown artifact type {artifact_type!r} (case {case_id})")
        if not pp_title:
            errors.append(f"missing PP Title (case {case_id})")
        if not controlled_question:
            errors.append(f"missing Controlled Question (case {case_id})")

    if errors:
        raise SystemExit("Frozen manifest failed projection invariants:\n  " + "\n  ".join(errors))


def main() -> None:
    if not _MANIFEST_PATH.is_file():
        raise SystemExit(
            f"frozen manifest not found at {_MANIFEST_PATH} — place the authoritative "
            f"{_SOURCE_MANIFEST_NAME} there before regenerating the projection"
        )

    manifest_bytes = _MANIFEST_PATH.read_bytes()
    source_sha256 = hashlib.sha256(manifest_bytes).hexdigest()

    workbook = openpyxl.load_workbook(_MANIFEST_PATH, data_only=True)
    worksheet = workbook[_SHEET_NAME]
    header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    case_id_col = header.index("Case ID")
    population_id_col = header.index("PP ID")
    artifact_type_col = header.index("Expected Primary Artifact Type")
    pp_title_col = header.index("PP Title")
    controlled_question_col = header.index("Controlled Question")

    cases: list[dict[str, str]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[case_id_col] is None:
            continue
        cases.append(
            {
                "case_id": row[case_id_col],
                "population_id": row[population_id_col],
                "expected_primary_artifact_type": row[artifact_type_col],
                "pp_title": row[pp_title_col],
                "controlled_question": row[controlled_question_col],
            }
        )

    _validate(cases)

    projection = {
        "source_manifest": _SOURCE_MANIFEST_NAME,
        "source_manifest_version": _SOURCE_MANIFEST_VERSION,
        "source_manifest_sha256": source_sha256,
        "generated_by": "08_Development/implementation/scripts/generate_evaluation_case_manifest_projection.py",
        "case_count": len(cases),
        "cases": cases,
    }

    _PROJECTION_PATH.write_text(json.dumps(projection, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(cases)} validated cases to {_PROJECTION_PATH} (source sha256={source_sha256})")


if __name__ == "__main__":
    main()
